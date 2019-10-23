
import json
import operator
import openpyxl as xl
import requests
from mock import Mock



def GET(url):
    if url=="http://localhost:3000/":
        global data, columnNames, modelDict, wrngColEx, parameters_count, target, wrong_target, path, libraries
        with open("data.json") as json_file:
            data = json.load(json_file)
        return data

    

# USECASE 1
def mockbestModel(path, target):
    temp_get = requests.get
    requests.get = Mock(side_effect=GET)
    data = requests.get(url = "http://localhost:3000/")
    columnNames = data["columnNames"]
    modelDict = data["listModels"]
    # for mocking the main and alternate flow
    flag = targetCheck(target, columnNames)
    if flag != 1:
        requests.get = temp_get
        return flag
    sorted_x = sorted(modelDict.items(), key=operator.itemgetter(1))
    requests.get = temp_get
    return str(sorted_x[len(sorted_x) - 1][0])

def targetCheck(target, columnNames):
    return 1 if target in columnNames else "The target column is not present in the file. Please upload the file again and give the correct target column name. Remember, target column is case sensitive."


# USECASE 2
def mock_analysis_interaction(path, target):
    temp_get = requests.get
    requests.get = Mock(side_effect=GET)
    data = requests.get(url="http://localhost:3000/")
    columnNames = data["columnNames"]
    parameters_count = int(data["parameters_to_be_counted"])
    flag = targetCheck(target, columnNames)
    if flag != 1:
        requests.get = temp_get
        return flag
    filename = data["path_for_usecase2"]
    a = ""
    count = 0
    for line in open(filename, 'r'):
        a = a + line
    if "MEAN" in a:
        count = count + 1
    if "MEDIAN" in a:
        count = count + 1
    if "MODE" in a:
        count = count + 1
    if "Correlation" in a:
        count = count + 1
    if "Normality Tests" in a:
        count = count + 1
    if count == parameters_count:
        requests.get = a
        return str(filename)
    requests.get = temp_get
    return 0


# USECASE 3
def mock_keyword_extraction(msg):
    temp_get = requests.get
    requests.get = Mock(side_effect=GET)
    data = requests.get(url="http://localhost:3000/")
    libraries = data["libraries"]

    list1 = str(msg).split(' ')
    ans_list = []
    length = len(list1)
    i = 0
    wb = xl.load_workbook(data["xlsx_file"])
    sheet = wb['Sheet1']
    libInfo = {}
    for row in range(2, sheet.max_row + 1):
        libInfo[sheet.cell(row, 1).value] = sheet.cell(row, 2).value
    while i < length:
        for r in (libInfo):
            if list1[i].lower() == r:
                ans_list.append({r: libraries[r]})
        i = i + 1
    requests.get = temp_get
    return ans_list
