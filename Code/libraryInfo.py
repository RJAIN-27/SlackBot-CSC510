import openpyxl as xl
import json
wb = xl.load_workbook('libraryFile.xlsx')
sheet = wb['Sheet1']
# requested library key word is passed
requestedLibrary = "keras"
requestedLibrary = requestedLibrary.lower()
libInfo = {}
#print(sheet.max_row)


for row in range(2 , sheet.max_row + 1):
    #print(sheet.cell(row, 1).value)
    libInfo[sheet.cell(row, 1).value] = sheet.cell(row, 2).value

# requested library details are forwarded
if(requestedLibrary in libInfo):
    jsobject = {requestedLibrary : libInfo[requestedLibrary]}
    print(json.dumps(jsobject))
#print(libInfo)
