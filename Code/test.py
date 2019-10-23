# Python code to demonstrate working of unittest using mock
import operator
import unittest
from mock import patch
import json
import openpyxl as xl
import KeywordExtraction
import modelSelection

with open("data.json") as json_file:
    data = json.load(json_file)
libraries = data["libraries"]
columnNames = data["columnNames"]
modelDict = data["listModels"]
wrngColEx = data["wrongTargetColumnException"]
parameters_to_be_counted = int(data["parameters_to_be_counted"])
target = data["target"]
wrong_target = data["wrong_target"]
path_to_csv_file_case1_case2 = data["path_to_csv_file_case1_case2"]
filename = data["path_for_usecase2"]

# USECASE 1
def mockbestModel(path, target):
    # for mocking the main and alternate flow
    flag = targetCheck(target, columnNames)
    if flag != 1:
        return flag
    sorted_x = sorted(modelDict.items(), key=operator.itemgetter(1))
    return (sorted_x[len(sorted_x) - 1][0])

def max_val_fun():
    ls = []
    for i in modelDict:
        ls.append(modelDict[i])
    return max(ls)

# USECASE 1 & 2
def targetCheck(target, columnNames):
    return 1 if target in columnNames else wrngColEx

# USECASE 2
def mock_analysis_interaction(path, target):
    flag = targetCheck(target, columnNames)
    if flag != 1:
        return flag

    a = ""
    count = 0
    f = open(filename,'r')
    for line in f:
        a = a + line
    if "MEAN" in a:
        count = count + 1
    if "MEDIAN" in a:
        count = count + 1
    if "MODE" in a:
        count = count + 1
    if "Correlation" in a:
        count = count + 1
    if "Normality Tests" in a:
        count = count + 1
    if count == parameters_to_be_counted:
        return filename
    return 0

# USECASE 3
def mock_keyword_extraction(a):
    list = a.split(' ')
    ans_list = []
    length = len(list)
    i = 0
    wb = xl.load_workbook(data["xlsx_file"])
    sheet = wb['Sheet1']
    libInfo = {}
    for row in range(2, sheet.max_row + 1):
        libInfo[sheet.cell(row, 1).value] = sheet.cell(row, 2).value
    while i < length:
        for r in (libInfo):
            if list[i].lower() in r:
                ans_list.append({r: libraries[r]})
        i = i + 1
    return ans_list

def read_from_json_and_test(a):
    list = a.split(' ')
    i = 0
    ans_list = []

    length = len(list)
    while i < length:
        for r in (libraries):
            if list[i].lower() == r:
                ans_list.append({r: libraries[r]})
        i = i + 1
    return ans_list

class TestStringMethods(unittest.TestCase):
    # usecase 1 - happy flow
    @patch('modelSelection.modelSelInteraction', side_effect=mockbestModel)
    def test_modelsel(self, modelSelInteraction):
        self.assertEqual(modelDict[modelSelInteraction(path_to_csv_file_case1_case2, target)], max_val_fun())
    # usecase 1 - alternate flow
    @patch('modelSelection.modelSelInteraction', side_effect=mockbestModel)
    def test_modelsel2(self, modelSelInteraction):
        self.assertEqual(modelSelInteraction(path_to_csv_file_case1_case2, wrong_target), wrngColEx)

    # usecase2 - happy flow
    @patch('analysis.analysisInteraction', side_effect=mock_analysis_interaction)
    def test_analysis1(self, analysisInteraction):
        self.assertEqual(analysisInteraction(path_to_csv_file_case1_case2, target), filename)
    # usecase2 - alternate flow
    @patch('analysis.analysisInteraction', side_effect=mock_analysis_interaction)
    def test_analysis2(self, analysisInteraction):
        self.assertEqual(analysisInteraction(path_to_csv_file_case1_case2, wrong_target), wrngColEx)
    
    # usecase 3 - happy flow
    @patch('KeywordExtraction.keywordExtraction', side_effect=mock_keyword_extraction)
    def test_strings_a(self, keywordExtraction):
        self.assertEqual(keywordExtraction("know numpy pandas"), read_from_json_and_test("know numpy pandas"))
    # usecase 3 - alternate flow
    @patch('KeywordExtraction.keywordExtraction', side_effect=mock_keyword_extraction)
    def test_strings_b(self, keywordExtraction):
        self.assertEqual(keywordExtraction("jon snow knows nothing"), [])

if __name__ == '__main__':
    unittest.main()
