import KeywordExtraction
import modelSelection
import analysis
#import mocking_infrastructure
import json
import unittest
import test
#from mock import Mock
import collections
import openpyxl as xl
import keywordex


with open("/home/ubuntu/CSC510-23/Code/data.json") as json_file:
    jsonData = json.load(json_file)

path = "my.csv"
flag=0

# KeywordExtraction.keywordExtraction = Mock(side_effect=mocking_infrastructure.mock_keyword_extraction)
# modelSelection.modelSelInteraction = Mock(side_effect=mocking_infrastructure.mockbestModel)
# analysis.analysisInteraction = Mock(side_effect = mocking_infrastructure.mock_analysis_interaction)


class Command(object):
    
    def handlecommand(self, user, command):
        global flag
        if "know" in command.lower() and "from" in command.lower():
            list=KeywordExtraction.keywordExtraction1(command, "Sheet1", "Sheet2", "Sheet3")
            list.append("onlyfunction")
            return list
        elif "know" in command.lower():
            list=KeywordExtraction.keywordExtraction(command, "Sheet1")
            list.append("onlylibrary")
            #list=mocking_infrastructure.mock_keyword_extraction(command)
            print (list) 
            return list
        elif any(word in command for word in jsonData["model_sel_words"]):
            return "Please give the csv file that you want to analyze or want a suggestion about"
        elif "Pssst" in command.lower():
            return "error"    
        elif "no" in command.lower():
            return "Sorry about that :("
        elif "yes" in command.lower():
            return "Thankyou for the feedback"
        elif flag==2:
            model_Selection=modelSelection.modelSelInteraction(path, command)
            flag=0
            #model_Selection=mocking_infrastructure.mockbestModel(path,command)
            return model_Selection
        elif flag==1:
            analysis_file=analysis.analysisInteraction(path, command)
            flag=0
            #analysis_file=mocking_infrastructure.mock_analysis_interaction(path,command)
            return analysis_file
        elif flag==0:
             return "Sorry I did not get you"
        
    def handlecommands(self, user, command):
        global flag
        command=command.lower()
        if any(word in command for word in jsonData["model_sel_words"]):
            flag=2
            response="Please provide the target column"
            return response

        if any(word in command for word in jsonData["analysis_words"]):
            flag=1
            response="Please provide the target column"
            return response
        
        else:
            return "Please give the file again with instructions on what to perform like suggesting model or analyzing the Data"

    def handlecommande(self, user, command, listx):
        ans=[]
        command=command.lower()
        if "know" in command:
            listi=KeywordExtraction.keywordExtraction(command, "Sheet3")
            d=collections.defaultdict(list)
            wb = xl.load_workbook(jsonData["xlsx_file"])
            sheet1= wb['Sheet2']
            for row in range(2, sheet1.max_row + 1):
                d[sheet1.cell(row, 1).value].append((sheet1.cell(row, 2).value))

            for i in listi:
                for j in listx:
                    m=d[j]
                    for k in i:
                        if k in m:
                            ans.append(i)
            return ans
        elif "no" in command:
            return "Please continue with either of the functionalities like knowing more about libraries, data analysis or suggestion"
             
        else:
            return "Sorry I did not get you"
