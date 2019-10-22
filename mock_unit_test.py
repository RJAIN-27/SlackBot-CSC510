# Python code to demonstrate working of unittest using mock
import operator
import unittest
from mock import patch
import json
import openpyxl as xl
import KeywordExtraction
import modelSelection

with open("data.json") as json_file:
    data = json.load(json_file)
columnNames = data["columnNames"]
modelDict = data["listModels"]
wrngColEx = data["wrongTargetColumnException"]

# USECASE 1
def mockbestModel(modelDict,target,columnNames):
    # for mocking the main and alternate flow
    flag = targetCheck(target,columnNames)
    if flag !=1:
        return flag
    sorted_x = sorted(modelDict.items(), key=operator.itemgetter(1))
    return(sorted_x[len(sorted_x)-1][0])

def max_val_fun():
    ls = []
    for i in modelDict:
        ls.append(modelDict[i])
    return max(ls)
# USECASE 2

# USECASE 1 & 2
def targetCheck(target, columnNames):
    return 1 if target in columnNames else "The target column is not present in the file. Please upload the file again and give the correct target column name. Remember, target column is case sensitive."

# USECASE 3
def mock_keyword_extraction(a):
    list = a.split(' ')
    ans_list = []
    length = len(list)
    i = 0
    wb = xl.load_workbook('libraryFile.xlsx')
    sheet = wb['Sheet1']
    libInfo = {}
    for row in range(2, sheet.max_row + 1):
        libInfo[sheet.cell(row, 1).value] = sheet.cell(row, 2).value
    while i < length:
        for r in (libInfo):
            if list[i] in r:
                ans_list.append(list[i])
        i = i + 1

def read_from_json_and_test(a):
    list = a.split(' ')
    i = 0
    ans_list = []
    length = len(list)
    while i < length:
        for r in (data):
            if list[i] in r:
                ans_list.append(list[i])
        i = i + 1

class TestStringMethods(unittest.TestCase):
    # usecase 1 - happy flow
    @patch('modelSelection.modelSelInteraction', side_effect=mockbestModel)
    def test_modelsel(self,modelSelInteraction):
        self.assertEqual(modelDict[modelSelInteraction(modelDict,"Class",columnNames)],max_val_fun())
    # usecase 1 - alternate flow
    @patch('modelSelection.modelSelInteraction', side_effect=mockbestModel)
    def test_modelsel2(self, modelSelInteraction):
        self.assertEqual(modelSelInteraction(modelDict, "class", columnNames), wrngColEx)

    # usecase 3 - happy flow
    @patch('KeywordExtraction.keywordExtraction', side_effect=mock_keyword_extraction)
    def test_strings_a(self, keywordExtraction):
        self.assertEqual(keywordExtraction("know numpy pandas"), read_from_json_and_test("know numpy pandas"))
    # usecase 3 - alternate flow
    @patch('KeywordExtraction.keywordExtraction', side_effect=mock_keyword_extraction)
    def test_strings_b(self, keywordExtraction):
        self.assertEqual(keywordExtraction("jon snow knows nothing"), None)
